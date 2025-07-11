import csv
import openpyxl
import logging
from datetime import datetime, date, time
from typing import List, Dict, Union, Optional, Any, cast
from openpyxl.worksheet.worksheet import Worksheet
#from mypy.typeshed.stdlib.decimal import Decimal # type: ignore


logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO)


def read_csv_transactions(file_path: str) -> List[Dict[str, Union[str, int, datetime]]]:
    """
    Чтение транзакций из CSV файла
    """
    transactions = []

    with open(file_path, mode='r', encoding='utf-8') as file:
        # Определяем разделитель как точку с запятой
        reader = csv.DictReader(file, delimiter=';')

        for row in reader:
            # Преобразуем строку даты в объект datetime
            try:
                row['date'] = datetime.strptime(row['date'], '%Y-%m-%dT%H:%M:%SZ')
            except ValueError:
                row['date'] = None

            # Преобразуем amount в число
            try:
                row['amount'] = int(row['amount'])
            except (ValueError, KeyError):
                row['amount'] = 0

            transactions.append(row)

    return transactions

def read_xlsx_transactions(
        file_path: str, sheet_name: Optional[str] = None
) -> List[Dict[str, Union[str, int, datetime, None]]]:

    transactions: List[Dict[str, Any]] = []

    workbook = openpyxl.load_workbook(file_path)

     # Получение листа
    sheet: Optional[Worksheet] = None
    if sheet_name:
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            logger.warning(f"Лист с именем '{sheet_name}' не найден в файле {file_path}.")
            return transactions  # Возвращаем пустой список
    else:
        sheet = workbook.active

    if sheet is None:
        logger.warning(f"Не удалось получить активный лист из файла {file_path}.")
        return transactions

    headers = [str(cell.value) if cell.value is not None else "" for cell in sheet[1]]

    for row in sheet.iter_rows(min_row=2, values_only=True):
        transaction = dict(zip(headers, row))

        if not transaction.get('id'):
            continue

        # Обработка даты
        date_val = transaction.get('date')
        if isinstance(date_val, str):
            try:
                transaction['date'] = datetime.strptime(date_val, '%Y-%m-%dT%H:%M:%SZ')
            except ValueError:
                logger.debug(f"Неверный формат даты: {date_val}")
                transaction['date'] = None
        elif not isinstance(date_val, datetime):
            transaction['date'] = None

        transactions.append(transaction)

    return transactions
# def read_xlsx_transactions(
#     file_path: str, sheet_name: Optional[str] = None
# ) -> List[Dict[str, Union[str, int, datetime, None]]]:
#
#     transactions = []
#
#     workbook = openpyxl.load_workbook(file_path)
#     sheet = workbook[sheet_name] if sheet_name else workbook.active
#
#     headers = [str(cell.value) if cell.value is not None else "" for cell in sheet[1]]
#
#     for row in sheet.iter_rows(min_row=2, values_only=True):
#         transaction = dict(zip(headers, row))
#
#         if not transaction.get('id'):
#             continue
#
#         # Обработка даты
#         date_val = transaction.get('date')
#         if isinstance(date_val, str):
#             try:
#                 transaction['date'] = datetime.strptime(date_val, '%Y-%m-%dT%H:%M:%SZ')
#             except ValueError:
#                 transaction['date'] = None
#         elif isinstance(date_val, datetime):
#             pass
#         else:
#             transaction['date'] = None

        # Обработка amount с явной проверкой типов
        # amount_val = transaction.get('amount')
        # if isinstance(amount_val, (int, float, Decimal, bool)):
        #     transaction['amount'] = int(amount_val)
        # elif isinstance(amount_val, str):
        #     try:
        #         transaction['amount'] = int(amount_val)
        #     except ValueError:
        #         transaction['amount'] = 0
        # else:
        #     transaction['amount'] = 0
# amount_val = transaction.get('amount')
#
# if isinstance(amount_val, (int, float, bool)):
#     transaction['amount'] = int(amount_val)
# elif isinstance(amount_val, str):
#     try:
#         transaction['amount'] = int(float(amount_val))  # сначала в float, затем в int
#     except ValueError:
#         transaction['amount'] = 0
# else:
#     transaction['amount'] = 0
#
#         # Приведение остальных значений к str | int | datetime | None
#         for key, val in transaction.items():
#             if key in ('date', 'amount'):
#                 continue  # уже обработаны
#             if val is None:
#                 transaction[key] = ""
#             elif isinstance(val, (str, int, datetime)):
#                 pass  # оставляем как есть
#             else:
#                 # Для прочих типов, например CellRichText, date, time и т.д. — преобразуем в строку
#                 transaction[key] = str(val)
#
#         transactions.append(transaction)
#
#     return transactions


def process_transactions(transactions: List[Dict]) -> None:
    """
    Обработка и вывод информации о транзакциях
    """
    print(f"Всего транзакций: {len(transactions)}")

    if not transactions:
        return

    # Статистика по статусам
    status_counts: Dict[str, int] = {}
    for t in transactions:
        status = t.get('state', 'UNKNOWN')
        status_counts[status] = status_counts.get(status, 0) + 1

    print("\nСтатистика по статусам:")
    for status, count in status_counts.items():
        print(f"{status}: {count}")

    # Топ-5 валют по количеству транзакций
    currency_counts: Dict[str, int] = {}
    for t in transactions:
        currency = t.get('currency_name', 'UNKNOWN')
        currency_counts[currency] = currency_counts.get(currency, 0) + 1

    top_currencies = sorted(currency_counts.items(), key=lambda x: x[1], reverse=True)[:5]
    print("\nТоп-5 валют по количеству транзакций:")
    for currency, count in top_currencies:
        print(f"{currency}: {count}")

    # Самая крупная транзакция
    max_amount = max(t['amount'] for t in transactions if 'amount' in t)
    largest_transactions = [t for t in transactions if t.get('amount') == max_amount]

    print("\nСамые крупные транзакции (сумма: {max_amount}):")
    for t in largest_transactions:
        print(f"ID: {t['id']}, Дата: {t['date']}, Валюта: {t['currency_name']}")


def main() -> None:
    # Обработка CSV файла
    print("Обработка CSV файла...")
    csv_transactions = read_csv_transactions('transactions.csv')
    process_transactions(csv_transactions)

    # Обработка XLSX файла
    print("\nОбработка XLSX файла...")
    xlsx_transactions = read_xlsx_transactions('transactions_excel.xlsx', 'Лист 1')
    process_transactions(xlsx_transactions)


if __name__ == "__main__":
    main()
