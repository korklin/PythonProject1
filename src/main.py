import csv
import json
import logging
from datetime import datetime
from typing import Any, Dict, List, Optional, Union

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from src.bank_utils import process_bank_operations, process_bank_search

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO)


def read_csv_transactions(file_path: str) -> List[Dict[str, Union[str, int, datetime]]]:
    transactions = []
    with open(file_path, mode="r", encoding="utf-8") as file:
        reader = csv.DictReader(file, delimiter=";")
        for row in reader:
            try:
                row["date"] = datetime.strptime(row["date"], "%Y-%m-%dT%H:%M:%SZ")
            except (ValueError, KeyError):
                row["date"] = None
            try:
                row["amount"] = int(row["amount"])
            except (ValueError, KeyError):
                row["amount"] = 0
            transactions.append(row)
    return transactions


def read_xlsx_transactions(file_path: str, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
    """Чтение транзакций из XLSX с минимальной строгой типизацией"""
    transactions = []

    try:
        # 1. Загрузка книги (без строгой типизации)
        workbook = openpyxl.load_workbook(file_path)

        # 2. Получение листа (прагматичный подход)
        sheet = None
        if sheet_name:
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]  # type: ignore
        if sheet is None:
            sheet = workbook.active

        # 3. Проверка что это действительно Worksheet
        if not isinstance(sheet, Worksheet):
            raise ValueError("Unsupported sheet type")

        # 4. Чтение данных (прагматичная обработка)
        headers = [str(cell.value) if cell.value else f"col_{i + 1}" for i, cell in enumerate(sheet[1])]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                transaction = dict(zip(headers, row))
                if not transaction.get("id"):
                    continue

                # Преобразование даты
                if "date" in transaction:
                    if isinstance(transaction["date"], str):
                        try:
                            transaction["date"] = datetime.strptime(transaction["date"], "%Y-%m-%dT%H:%M:%SZ")
                        except ValueError:
                            transaction["date"] = None
                    elif not isinstance(transaction["date"], datetime):
                        transaction["date"] = None

                transactions.append(transaction)
            except Exception as row_error:
                logger.warning(f"Row processing error: {row_error}")
                continue

    except Exception as e:
        logger.error(f"XLSX read failed: {e}")
        raise

    return transactions


def read_json_transactions(file_path: str) -> Any:
    with open(file_path, "r", encoding="utf-8") as file:
        return json.load(file)


def filter_transactions_by_status(transactions: List[Dict], status: str) -> List[Dict]:
    status_clean = status.strip().upper()
    return [t for t in transactions if t.get("state", "").strip().upper() == status_clean]


def sort_transactions(transactions: List[Dict], ascending: bool = True) -> List[Dict]:
    return sorted(transactions, key=lambda t: t.get("date") or datetime.min, reverse=not ascending)


def filter_rub(transactions: List[Dict]) -> List[Dict]:
    return [t for t in transactions if t.get("currency_name", "").strip().lower() == "руб."]


def print_transaction(t: Dict) -> None:
    date = t.get("date")
    if isinstance(date, datetime):
        date_str = date.strftime("%d.%m.%Y")
    else:
        date_str = "неизвестно"
    description = t.get("description", "Без описания")
    from_ = t.get("from", "")
    to_ = t.get("to", "")
    amount = t.get("amount", 0)
    currency = t.get("currency_name", "")
    print(f"\n{date_str} {description}")
    if from_ and to_:
        print(f"{from_} -> {to_}")
    elif to_:
        print(f"Счет -> {to_}")
    elif from_:
        print(f"{from_} -> Счет")
    print(f"Сумма: {amount} {currency}")


def process_transactions(transactions: List[Dict]) -> None:
    if not transactions:
        print("Список транзакций пуст.")
        return

    # --- фильтрация по статусу ---
    available_statuses = {"EXECUTED", "CANCELED", "PENDING"}
    while True:
        status_input = input("\nВведите статус для фильтрации (EXECUTED, CANCELED, PENDING): ").strip().upper()
        if status_input in available_statuses:
            transactions = filter_transactions_by_status(transactions, status_input)
            print(f"Операции отфильтрованы по статусу: {status_input}")
            break
        else:
            print(f"Статус '{status_input}' недопустим.")

    if not transactions:
        print("Не найдено ни одной транзакции с указанным статусом.")
        return

    # --- сортировка по дате ---
    if input("Отсортировать операции по дате? Да/Нет: ").strip().lower() == "да":
        direction = input("По возрастанию или по убыванию?: ").strip().lower()
        transactions = sort_transactions(transactions, ascending=("возрастанию" in direction))

    # --- только рублевые ---
    if input("Выводить только рублевые транзакции? Да/Нет: ").strip().lower() == "да":
        transactions = filter_rub(transactions)

    # --- поиск по описанию ---
    if input("Отфильтровать по слову в описании? Да/Нет: ").strip().lower() == "да":
        keyword = input("Введите ключевое слово: ").strip()
        transactions = process_bank_search(transactions, keyword)

    if not transactions:
        print("Не найдено ни одной транзакции, подходящей под ваши условия фильтрации.")
        return

    print("\nРаспечатываю итоговый список транзакций...")
    print(f"\nВсего банковских операций в выборке: {len(transactions)}")
    for t in transactions:
        print_transaction(t)

    # --- подсчет категорий ---
    if input("\nПодсчитать транзакции по категориям? Да/Нет: ").strip().lower() == "да":
        cats = input("Введите категории через запятую: ").split(",")
        categories = [c.strip() for c in cats if c.strip()]
        counts = process_bank_operations(transactions, categories)
        print("\nОперации по категориям:")
        for cat, count in counts.items():
            print(f"{cat}: {count}")


def main() -> None:
    print("Привет! Добро пожаловать в программу работы с банковскими транзакциями.")
    print("Выберите необходимый пункт меню:")
    print("1. Получить информацию о транзакциях из JSON-файла")
    print("2. Получить информацию о транзакциях из CSV-файла")
    print("3. Получить информацию о транзакциях из XLSX-файла")

    choice = input("Ваш выбор: ").strip()

    try:
        if choice == "1":
            path = input("Введите путь к JSON-файлу: ").strip()
            transactions = read_json_transactions(path)
            print("\nДанные из JSON-файла загружены.")
        elif choice == "2":
            path = input("Введите путь к CSV-файлу: ").strip()
            transactions = read_csv_transactions(path)
            print("\nДанные из CSV-файла загружены.")
        elif choice == "3":
            path = input("Введите путь к XLSX-файлу: ").strip()
            sheet = input("Введите имя листа (или нажмите Enter для первого): ").strip() or None
            transactions = read_xlsx_transactions(path, sheet)
            print("\nДанные из XLSX-файла загружены.")
        else:
            print("Некорректный выбор.")
            return
    except Exception as e:
        print(f"Ошибка при загрузке файла: {e}")
        return

    process_transactions(transactions)


if __name__ == "__main__":
    main()
